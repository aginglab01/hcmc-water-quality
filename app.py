"""
호치민 수질 모니터링 대시보드
Streamlit + Folium
"""

import re
import random
import openpyxl
import pandas as pd
import folium
from folium.plugins import MarkerCluster
import streamlit as st
from streamlit_folium import st_folium
from collections import defaultdict

# ── 페이지 설정 ───────────────────────────────────────────────────────
st.set_page_config(
    page_title="호치민 수질 모니터링",
    page_icon="🚰",
    layout="wide",
)

# ── 구 좌표 테이블 ────────────────────────────────────────────────────
DISTRICT_COORDS = {
    "Q.1":  (10.7769, 106.7009), "Q.2":  (10.7873, 106.7491),
    "Q.3":  (10.7769, 106.6839), "Q.4":  (10.7538, 106.7025),
    "Q.5":  (10.7520, 106.6683), "Q.6":  (10.7452, 106.6325),
    "Q.7":  (10.7252, 106.7213), "Q.8":  (10.7258, 106.6594),
    "Q.9":  (10.8327, 106.8018), "Q.10": (10.7727, 106.6691),
    "Q.11": (10.7652, 106.6505), "Q.12": (10.8693, 106.6465),
    "Q. Binh Thanh": (10.8123, 106.7091),
    "Q. Go Vap":     (10.8384, 106.6649),
    "Q. Phu Nhuan":  (10.7993, 106.6793),
    "Q. Tan Binh":   (10.8024, 106.6524),
    "Q. Tan Phu":    (10.7908, 106.6273),
    "Q. Binh Tan":   (10.7656, 106.6022),
    "Q. Thu Duc":    (10.8527, 106.7539),
    "H. Nha Be":     (10.6683, 106.7183),
    "H. Hoc Mon":    (10.8883, 106.5948),
    "H. Binh Chanh": (10.6833, 106.5833),
    "H. Cu Chi":     (11.0017, 106.4932),
    "H. Can Gio":    (10.4076, 106.9510),
}

# 악센트 제거 매핑
ACCENT_MAP = str.maketrans(
    "àáảãạăắằẳẵặâấầẩẫậèéẻẽẹêếềểễệìíỉĩịòóỏõọôốồổỗộơớờởỡợùúủũụưứừửữựỳýỷỹỵđ"
    "ÀÁẢÃẠĂẮẰẲẴẶÂẤẦẨẪẬÈÉẺẼẸÊẾỀỂỄỆÌÍỈĨỊÒÓỎÕỌÔỐỒỔỖỘƠỚỜỞỠỢÙÚỦŨỤƯỨỪỬỮỰỲÝỶỸỴĐ",
    "aaaaaaaaaaaaaaaaaeeeeeeeeeeeiiiiiooooooooooooooooouuuuuuuuuuuyyyyyd"
    "AAAAAAAAAAAAAAAAAEEEEEEEEEEEIIIIIOOOOOOOOOOOOOOOOOUUUUUUUUUUUYYYYYD"
)

def normalize(text: str) -> str:
    return text.translate(ACCENT_MAP).lower()

DISTRICT_NORMALIZED = {normalize(k): v for k, v in DISTRICT_COORDS.items()}

def extract_coords(addr: str):
    if not addr:
        return None
    addr_n = normalize(addr)
    # 긴 키부터 매칭
    for key_n, coords in sorted(DISTRICT_NORMALIZED.items(), key=lambda x: -len(x[0])):
        if key_n in addr_n:
            return coords
    # 숫자 구 번호 매칭 (Q.X, 0.X OCR 오류 포함)
    m = re.search(r'[qQ0oO]\.?\s*(\d{1,2})', addr)
    if m:
        k = f"Q.{m.group(1)}"
        if k in DISTRICT_COORDS:
            return DISTRICT_COORDS[k]
    return None

# ── 데이터 로딩 (캐시) ────────────────────────────────────────────────
@st.cache_data
def load_data(path="water_quality_data_preview.xlsx"):
    wb = openpyxl.load_workbook(path)
    ws = wb["수질데이터"]
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    df = pd.DataFrame(rows)
    df["취수원주소"] = df["취수원주소"].fillna("").str.strip()
    df = df[df["취수원주소"] != ""]

    # 좌표 부여 (주소별 1회 계산 + 소량 jitter)
    rng = random.Random(42)
    coord_cache = {}
    def get_coords(addr):
        if addr not in coord_cache:
            c = extract_coords(addr)
            if c:
                coord_cache[addr] = (
                    c[0] + rng.uniform(-0.008, 0.008),
                    c[1] + rng.uniform(-0.008, 0.008),
                )
            else:
                coord_cache[addr] = None
        return coord_cache[addr]

    df["coords"] = df["취수원주소"].map(get_coords)
    df["lat"] = df["coords"].map(lambda c: c[0] if c else None)
    df["lon"] = df["coords"].map(lambda c: c[1] if c else None)
    df["연도"] = pd.to_numeric(df["연도"], errors="coerce")
    return df

# ── 집계 함수 ─────────────────────────────────────────────────────────
def agg_by_address(df):
    """주소별 집계 → 지도 핀 1개당 1행"""
    records = []
    for addr, grp in df.groupby("취수원주소"):
        lat = grp["lat"].dropna().iloc[0] if grp["lat"].notna().any() else None
        lon = grp["lon"].dropna().iloc[0] if grp["lon"].notna().any() else None
        if lat is None:
            continue

        statuses = grp["적합여부"].dropna().astype(str).tolist()
        if any("부" in s or "fail" in s.lower() for s in statuses):
            color = "red"
            status_label = "부적합 포함"
        elif statuses:
            color = "green"
            status_label = "적합"
        else:
            color = "gray"
            status_label = "미기재"

        def avg(col):
            v = pd.to_numeric(grp[col], errors="coerce").dropna()
            return round(v.mean(), 3) if len(v) else "N/A"

        years = sorted(grp["연도"].dropna().astype(int).unique())
        districts = ", ".join(sorted(grp["동(Phường)"].dropna().astype(str).unique()))
        records.append({
            "addr": addr, "lat": lat, "lon": lon,
            "color": color, "status_label": status_label,
            "count": len(grp),
            "years": ", ".join(str(y) for y in years),
            "districts": districts,
            "avg_pH": avg("pH"),
            "avg_turbidity": avg("탁도(NTU)"),
            "avg_chlorine": avg("잔류염소(mg/L)"),
        })
    return records

# ── 지도 생성 ─────────────────────────────────────────────────────────
def build_map(pin_records):
    m = folium.Map(location=[10.8231, 106.6297], zoom_start=11)
    cluster = MarkerCluster(
        options={"maxClusterRadius": 40, "disableClusteringAtZoom": 15}
    ).add_to(m)

    for r in pin_records:
        status_html = (
            '<span style="color:red"><b>⚠ 부적합 포함</b></span>' if r["color"] == "red"
            else '<span style="color:green">✓ 적합</span>' if r["color"] == "green"
            else "미기재"
        )
        popup_html = f"""
        <div style="font-family:Arial;font-size:13px;min-width:230px">
          <b>🚰 {r['addr']}</b>
          <hr style="margin:5px 0">
          <b>동(Phường):</b> {r['districts'] or '-'}<br>
          <b>측정 횟수:</b> {r['count']}회 &nbsp; <b>연도:</b> {r['years']}<br>
          <hr style="margin:5px 0">
          <b>평균 pH:</b> {r['avg_pH']}<br>
          <b>평균 탁도:</b> {r['avg_turbidity']} NTU<br>
          <b>평균 잔류염소:</b> {r['avg_chlorine']} mg/L<br>
          <b>적합여부:</b> {status_html}
        </div>"""
        folium.Marker(
            location=[r["lat"], r["lon"]],
            popup=folium.Popup(popup_html, max_width=290),
            tooltip=r["addr"][:45] + ("…" if len(r["addr"]) > 45 else ""),
            icon=folium.Icon(color=r["color"], icon="tint", prefix="fa"),
        ).add_to(cluster)
    return m

# ════════════════════════════════════════════════════════════════════════
# UI
# ════════════════════════════════════════════════════════════════════════
st.title("🚰 호치민시 수질 모니터링 대시보드")
st.caption("Ho Chi Minh City Water Quality Monitoring — 취수원별 현장 측정 데이터")

with st.spinner("데이터 로딩 중..."):
    df = load_data()

# ── 사이드바 필터 ─────────────────────────────────────────────────────
with st.sidebar:
    st.header("🔍 필터")

    years = sorted(df["연도"].dropna().astype(int).unique())
    sel_years = st.multiselect("연도", years, default=years)

    # 구(Quận) 추출
    def extract_district_label(addr):
        for key in sorted(DISTRICT_COORDS, key=len, reverse=True):
            if normalize(key) in normalize(addr):
                return key
        m = re.search(r'[qQ]\.?\s*(\d{1,2})', addr)
        if m:
            return f"Q.{m.group(1)}"
        return "기타"

    df["구"] = df["취수원주소"].map(extract_district_label)
    all_districts = sorted(df["구"].unique())
    sel_districts = st.multiselect("구(Quận/Huyện)", all_districts, default=all_districts)

    status_options = ["전체", "적합만", "부적합 포함", "미기재"]
    sel_status = st.radio("적합 여부", status_options)

    st.divider()
    st.caption("📍 핀 위치는 구(Quận) 단위 근사치입니다")

# ── 필터 적용 ─────────────────────────────────────────────────────────
filtered = df.copy()
if sel_years:
    filtered = filtered[filtered["연도"].isin(sel_years)]
if sel_districts:
    filtered = filtered[filtered["구"].isin(sel_districts)]

# 적합여부 필터
def status_group(row):
    s = str(row.get("적합여부") or "")
    if "부" in s or "fail" in s.lower():
        return "부적합 포함"
    elif s.strip():
        return "적합만"
    return "미기재"

if sel_status != "전체":
    filtered = filtered[filtered.apply(status_group, axis=1) == sel_status]

# ── 요약 통계 ─────────────────────────────────────────────────────────
total = len(filtered)
located = filtered["lat"].notna().sum()
unique_addr = filtered["취수원주소"].nunique()

col1, col2, col3, col4 = st.columns(4)
col1.metric("총 측정 건수", f"{total:,}건")
col2.metric("고유 측정소", f"{unique_addr:,}개")
col3.metric("지도 표시 가능", f"{filtered[filtered['lat'].notna()]['취수원주소'].nunique():,}개")

ph_vals = pd.to_numeric(filtered["pH"], errors="coerce").dropna()
col4.metric("평균 pH", f"{ph_vals.mean():.2f}" if len(ph_vals) else "N/A")

st.divider()

# ── 지도 ─────────────────────────────────────────────────────────────
col_map, col_info = st.columns([3, 1])

with col_map:
    st.subheader("📍 측정소 지도")
    if total == 0:
        st.warning("선택한 필터에 해당하는 데이터가 없습니다.")
    else:
        pin_records = agg_by_address(filtered)
        m = build_map(pin_records)
        st_folium(m, width=None, height=520, returned_objects=[])

with col_info:
    st.subheader("📊 현황")

    # 적합여부 분포
    status_counts = {"적합": 0, "부적합": 0, "미기재": 0}
    for _, row in filtered.iterrows():
        s = str(row.get("적합여부") or "")
        if "부" in s or "fail" in s.lower():
            status_counts["부적합"] += 1
        elif s.strip():
            status_counts["적합"] += 1
        else:
            status_counts["미기재"] += 1

    st.write("**적합여부 분포**")
    if total > 0:
        for label, cnt in status_counts.items():
            pct = cnt / total * 100
            color = "🔴" if label == "부적합" else "🟢" if label == "적합" else "⚫"
            st.write(f"{color} {label}: **{cnt:,}건** ({pct:.1f}%)")

    st.divider()

    # 구별 측정 건수 Top 10
    st.write("**구별 측정 건수 (Top 10)**")
    district_counts = filtered["구"].value_counts().head(10)
    st.dataframe(
        district_counts.reset_index().rename(columns={"구": "구(Quận)", "count": "건수"}),
        hide_index=True,
        use_container_width=True,
    )

# ── 원본 데이터 테이블 ────────────────────────────────────────────────
with st.expander("📋 원본 데이터 보기"):
    display_cols = ["취수원주소", "동(Phường)", "연도", "월", "pH",
                    "탁도(NTU)", "잔류염소(mg/L)", "Coliform(CFU/100mL)",
                    "E.coli(CFU/100mL)", "적합여부"]
    st.dataframe(filtered[display_cols].reset_index(drop=True), use_container_width=True)
